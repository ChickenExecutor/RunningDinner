import openpyxl
import os

class RunningDinnerHelper:
    def __init__(self, path):
        self.path = path
        workbook = openpyxl.load_workbook(self.path)
        self.sheet = workbook.active
        self.max_row_length: int = 0
        self.row_vorspeise: int = self.find_cell_in_column(self.sheet, 1, "Vorspeise")
        self.row_hauptgang: int = self.find_cell_in_column(self.sheet, 1, "Hauptgang")
        self.row_dessert: int = self.find_cell_in_column(self.sheet, 1, "Dessert")
        self.end_row: int = self.find_end()
        self.find_max_row_length()
        self.teams = self.create_teams()
        self.find_row_with_team()
        self.max_team_length = self.find_max_team_length()        

    def find_max_team_length(self):
        max_length: int = 0
        for team in self.teams:
            if len(team) > max_length:
                max_length = len(team)
        return max_length

    def clear_paste_area(self, max_team_length):
        for row in range(self.row_vorspeise + 1, self.end_row + 1):
            for col in range(self.max_row_length + 3, self.max_row_length + 3 +max_team_length + 1):
                self.sheet.cell(row=row, column=col).value = None
                self.sheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(fill_type=None)

    def find_cell_in_column(self, sheet, column_number, search_value):
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=column_number).value
            if cell_value == search_value:
                return row
        return None
    

    def find_end(self):
        for row in range(self.row_dessert, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row=row, column=1).value
            if cell_value is None:
                return row
        return self.sheet.max_row
    
    def create_teams(self):
        teams: list = []
        gang: str = "Vorspeise"
        for row in range(self.row_vorspeise + 1, self.end_row+1):
            cell_value = self.sheet.cell(row=row, column=1).value
            if cell_value == "Hauptgang":
                gang = "Hauptgang"
            elif cell_value == "Dessert":
                gang = "Dessert"
            elif cell_value != None:
                teams.append([gang, cell_value])
        return teams
    
    def find_max_row_length(self):
        for row in range(self.row_vorspeise, self.end_row+1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if cell_value != None:
                    if col > self.max_row_length:
                        self.max_row_length = col
                else:
                    break

    def find_row_with_team(self):
        for team in self.teams:
            for row in range(self.row_vorspeise+1, self.end_row + 1):
                for col in range(1, self.max_row_length + 1):
                    cell_value = self.sheet.cell(row=row, column=col).value
                    if cell_value == team[1]:
                        for col in range(1,self.max_row_length + 1):
                            cell_value = self.sheet.cell(row=row, column=col).value
                            if cell_value != team[1]:
                                team.append(cell_value)
                        break

    def get_team_colour(self, team_name):
        for row in range(self.row_vorspeise+1, self.end_row + 1):
            cell_value = self.sheet.cell(row=row, column=1).value
            if cell_value == team_name:
                team_colour =  self.sheet.cell(row=row, column=1).fill.start_color
                # if team_colour.type == 'rgb':
                #     return team_colour.rgb
                # if team_colour.type == 'indexed':
                #     # convert indexed theme/index color to ARGB (use openpyxl's COLOR_INDEX)
                #     team_colour = (('FF' + openpyxl.styles.colors.COLOR_INDEX[int(team_colour)]) if len(openpyxl.styles.colors.COLOR_INDEX[int(team_colour)]) == 6 else openpyxl.styles.colors.COLOR_INDEX[int(team_colour)])
                # if team_colour.type == 'theme':
                #     theme_index = team_colour.theme
                #     tint = team_colour.tint
                #     rgb = openpyxl.styles.colors.THEME_COLORS[theme_index]
                #     if tint is not None and tint != 0:
                #         # Apply tint to rgb
                #         def apply_tint(value, tint):
                #             value = int(value, 16)
                #             if tint < 0:
                #                 value = int(value * (1 + tint))
                #             else:
                #                 value = int(value + (255 - value) * tint)
                #             return max(0, min(255, value))
                #         r = apply_tint(rgb[0:2], tint)
                #         g = apply_tint(rgb[2:4], tint)
                #         b = apply_tint(rgb[4:6], tint)
                #         rgb = f"{r:02X}{g:02X}{b:02X}"
                #     team_colour = 'FF' + rgb
                return team_colour
                

    def paste_pairing(self, teams):
        i: int = 0
        hauptgang_added: bool = False
        Dessert_added: bool = False
        for team in teams:
            j: int = 0
            for pairing in team:
                if pairing == "Vorspeise":
                    pass
                elif pairing == "Hauptgang":
                    if not hauptgang_added:
                        hauptgang_added = True
                        i += 2
                elif pairing == "Dessert":
                    if not Dessert_added:
                        Dessert_added = True
                        i += 2
                else:
                    self.sheet.cell(row=self.row_vorspeise+1+i, column=self.max_row_length + 3 + j).value = pairing
                    fill_color = self.get_team_colour(pairing)
                    self.sheet.cell(row=self.row_vorspeise+1+i, column=self.max_row_length + 3 + j).fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                j += 1
            if len(team) != len(set(team)):
                self.sheet.cell(row=self.row_vorspeise+1+i, column=self.max_row_length + 3 + j).value = "Doppelte Teamzuordnung!"
                self.sheet.cell(row=self.row_vorspeise+1+i, column=self.max_row_length + 3 + j).fill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type="solid")
            i += 1




if __name__ == "__main__":
    helper = RunningDinnerHelper()
    helper.clear_paste_area(helper.max_team_length)
    helper.paste_pairing(helper.teams)
    helper.sheet.parent.save(os.path.join(os.path.dirname(__file__), "RunningDinner.xlsx"))

    




